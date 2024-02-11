import openpyxl
import tkinter as tk
from tkinter import ttk
from ttkthemes import ThemedTk
from tkinter import messagebox, filedialog
from tkinter import Toplevel
import hashlib
import datetime
import mysql.connector


###################################### Wyświetlanie ekranów #####################################################


def show_login_screen():

    def hash_password(haslo):
        hash = hashlib.sha256(haslo.encode()).hexdigest()
        return hash

    logged_in_user = ""
    user_id = ""

    def set_logged_in_user(username):
        global logged_in_user
        logged_in_user = username

    def set_user_id(id):
        global user_id
        user_id = id

    def check_login():
        email = email_entry.get()
        password = password_entry.get()

        conn = mysql.connector.connect(
            host="db4free.net",
            user="magazyn",
            password="Inzynierka",
            database="magazyn_pd"
        )
        cursor = conn.cursor()

        cursor.execute('SELECT * FROM uzytkownicy WHERE email=%s', (email,))

        user = cursor.fetchone()

        if user:
            password_in_datebase = user[5]
            hash = hash_password(password)
            if password_in_datebase == hash:
                name = user[1]
                lastname = user[2]
                level = user[3]
                u_id = user[6]

                full_name = f"{name} {lastname}"

                set_logged_in_user(full_name)
                set_user_id(u_id)

                if level == "admin":
                    root.destroy()
                    show_admin_screen()
                elif level == "pracownik":
                    root.destroy()
                    show_user_screen()
                elif level == "magazyn":
                    root.destroy()
                    show_storage_screen()
            else:
                result_label.config(text="Błąd logowania")
        else:
            result_label.config(text="Błąd logowania")

    root = ThemedTk(theme="arc")
    root.title("Logowanie")

    frame = ttk.Frame(root, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    email_label = ttk.Label(frame, text="Email:")
    email_label.grid(row=0, column=0, sticky="w", pady=(20, 5))
    email_entry = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
    email_entry.grid(row=0, column=1, padx=5, pady=(20, 5), ipady=10)

    password_label = ttk.Label(frame, text="Hasło:")
    password_label.grid(row=1, column=0, sticky="w", pady=(5, 10))
    password_entry = ttk.Entry(frame, show="*", width=50, font=('Arial', 12), justify='left')
    password_entry.grid(row=1, column=1, padx=5, pady=(5, 10), ipady=10)

    login_button = ttk.Button(frame, text="Zaloguj", command=check_login)
    login_button.grid(row=2, column=0, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)

    result_label = ttk.Label(frame, text="", foreground="red")
    result_label.grid(row=3, column=0, columnspan=2, pady=(15, 20))

    root.resizable(width=False, height=False)

    root.mainloop()


def show_admin_screen():
    global users_tree

    root = ThemedTk(theme="arc")
    root.title("Ekran Admina")

    main_frame = ttk.Frame(root, padding=10)
    main_frame.grid(row=0, column=20, sticky="nsew")
    logged_in_label = ttk.Label(main_frame, text=f"Jesteś zalogowany jako: {logged_in_user}")
    logged_in_label.grid(row=1, column=0, pady=(0, 10), padx=(20, 0), sticky="ew")

    logout_button = ttk.Button(main_frame, text="Wyloguj", command=lambda: on_close(root), width=15)
    logout_button.grid(row=1, column=19, pady=(0, 10), padx=(0, 20), sticky="e", ipadx=10, ipady=5)

    change_password_button = ttk.Button(main_frame, text="Zmień hasło", command=change_password, width=15)
    change_password_button.grid(row=1, column=18, pady=(0, 10), sticky="e", ipadx=10, ipady=5)

    users_frame = ttk.Frame(main_frame, padding=10)
    users_frame.grid(row=0, column=0, columnspan=20, padx=10, pady=10, sticky="nsew")

    users_button = ttk.Button(users_frame, text="Pokaż użytkowników", command=populate_tree3)
    users_button.grid(row=0, column=0, pady=(0, 10), sticky="ew", ipadx=20, ipady=10)

    users_tree = ttk.Treeview(users_frame)
    users_tree["columns"] = ("ID", "Imie", "Nazwisko", "Rola", "Email", "Haslo", "Identyfikator")
    users_tree.heading("ID", text="ID", anchor='center')
    users_tree.heading("Imie", text="Imie", anchor='center')
    users_tree.heading("Nazwisko", text="Nazwisko", anchor='center')
    users_tree.heading("Rola", text="Rola", anchor='center')
    users_tree.heading("Email", text="Email", anchor='center')
    users_tree.heading("Haslo", text="Haslo", anchor='center')
    users_tree.heading("Identyfikator", text="Identyfikator", anchor='center')

    users_tree.column("#0", width=0, minwidth=0, stretch=tk.NO)
    users_tree.column("ID", width=150, stretch=False)
    users_tree.column("Imie", width=150)
    users_tree.column("Nazwisko", width=150)
    users_tree.column("Rola", width=150)
    users_tree.column("Email", width=150)
    users_tree.column("Haslo", width=150)
    users_tree.column("Identyfikator", width=150)
    users_tree.grid(row=2, column=0, pady=(0, 10), sticky="nsew")

    create_button = ttk.Button(users_frame, text="Utwórz użytkownika", command=create)
    create_button.grid(row=3, column=0, pady=(0, 10), sticky="ew")

    create_button = ttk.Button(users_frame, text="Edytuj użytkownika", command=edit)
    create_button.grid(row=4, column=0, pady=(0, 10), sticky="ew")

    root.resizable(width=False, height=False)

    root.mainloop()


# Ekran pracownika

def show_user_screen():
    global tree
    global order_tree
    global order__tree
    global root

    root = ThemedTk(theme="arc")
    root.title("Magazyn")

    main_frame = ttk.Frame(root, padding=10)
    main_frame.grid(row=0, column=20, sticky="nsew")
    logged_in_label = ttk.Label(main_frame, text=f"Jesteś zalogowany jako: {logged_in_user}")
    logged_in_label.grid(row=1, column=0, pady=(0, 10), padx=(20, 0), sticky="ew")

    logout_button = ttk.Button(main_frame, text="Wyloguj", command=lambda: on_close(root), width=15)
    logout_button.grid(row=1, column=19, pady=(0, 10), padx=(0, 20), sticky="e", ipadx=10, ipady=5)

    change_password_button = ttk.Button(main_frame, text="Zmień hasło", command=change_password, width=15)
    change_password_button.grid(row=1, column=18, pady=(0, 10), sticky="e", ipadx=10, ipady=5)

    inventory_frame = ttk.Frame(main_frame, padding=10)
    inventory_frame.grid(row=0, column=0, columnspan=10, padx=10, pady=10, sticky="nsew")

    data_entry_button = ttk.Button(inventory_frame, text="Pokaż stan magazynowy", command=populate_tree)
    data_entry_button.grid(row=0, column=0, pady=(0, 10), sticky="ew")

    tree = ttk.Treeview(inventory_frame)
    tree["columns"] = ("ID", "Nazwa", "Ilosc")
    tree.heading("ID", text="ID", anchor='center', command=lambda: sort_treeview(tree, "ID", False))
    tree.heading("Nazwa", text="Nazwa", anchor='center', command=lambda: sort_treeview2(tree, "Nazwa", False))
    tree.heading("Ilosc", text="Ilość", anchor='center', command=lambda: sort_treeview(tree, "Ilosc", False))
    tree.column("#0", width=0, minwidth=0, stretch=tk.NO)
    tree.column("ID", width=150, stretch=False)
    tree.column("Nazwa", width=150)
    tree.column("Ilosc", width=150)
    tree.grid(row=2, column=0, pady=(0, 10), sticky="nsew")

    add_to_order_button = ttk.Button(inventory_frame, text="Dodaj do zlecenia", command=add_to_order)
    add_to_order_button.grid(row=3, column=0, pady=(0, 10), sticky="ew", ipadx=20, ipady=10)

    order_frame = ttk.Frame(main_frame, padding=10)
    order_frame.grid(row=0, column=10, columnspan=10, padx=10, pady=10, sticky="nsew")

    button = ttk.Button(inventory_frame, text="Wyszukaj", command=search)
    button.grid(row=1, column=0, pady=(0, 10), sticky="ew")

    order_tree = ttk.Treeview(order_frame)
    order_tree["columns"] = ("ID", "Nazwa", "Ilosc", "Numer")
    order_tree.heading("ID", text="ID", anchor='center')
    order_tree.heading("Nazwa", text="Nazwa", anchor='center')
    order_tree.heading("Ilosc", text="Ilość", anchor='center')
    order_tree.heading("Numer", text="Numer Zlecenia", anchor='center')
    order_tree.column("#0", width=0, minwidth=0, stretch=tk.NO)
    order_tree.column("ID", width=100, stretch=False)
    order_tree.column("Nazwa", width=150)
    order_tree.column("Ilosc", width=100)
    order_tree.column("Numer", width=100, stretch=False)
    order_tree.grid(row=1, column=0, pady=(0, 10), columnspan=2, sticky="nsew")

    remove_from_order_button = ttk.Button(order_frame, text="Usuń zaznaczony", command=remove_from_order)
    remove_from_order_button.grid(row=2, column=0, pady=(0, 10), sticky="ew", columnspan=2)

    clear_order_button = ttk.Button(order_frame, text="Wyczyść listę", command=clear_window_without_send)
    clear_order_button.grid(row=3, column=0, pady=(0, 10), columnspan=2, sticky="ew")

    save_to_db_button = ttk.Button(order_frame, text="Wyślij zlecenie", command=save_to_database)
    save_to_db_button.grid(row=4, column=0, pady=(0, 10), columnspan=2, sticky="ew", ipadx=20, ipady=10)

    root.resizable(width=False, height=False)

    root.mainloop()


# Ekran magazyniera

def show_storage_screen():
    global tree
    global order_tree
    global order__tree
    global root

    root = ThemedTk(theme="arc")
    root.title("Magazyn")

    main_frame = ttk.Frame(root, padding=10)
    main_frame.grid(row=0, column=20, sticky="nsew")
    logged_in_label = ttk.Label(main_frame, text=f"Jesteś zalogowany jako: {logged_in_user}")
    logged_in_label.grid(row=1, column=0, pady=(0, 10), padx=(20, 0), sticky="ew")

    logout_button = ttk.Button(main_frame, text="Wyloguj", command=lambda: on_close(root), width=15)
    logout_button.grid(row=1, column=19, pady=(0, 10), padx=(0, 20), sticky="e", ipadx=10, ipady=5)

    change_password_button = ttk.Button(main_frame, text="Zmień hasło", command=change_password, width=15)
    change_password_button.grid(row=1, column=18, pady=(0, 10), sticky="e", ipadx=10, ipady=5)

    order_frame = ttk.Frame(main_frame, padding=10)
    order_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

    empty_row = ttk.Label(order_frame, text="")
    empty_row.grid(row=0, column=0, pady=(0, 40))

    data_add_button = ttk.Button(order_frame, text="Dodaj produkt", command=add_to_datebase)
    data_add_button.grid(row=1, column=0, pady=(0, 10), sticky="ew")

    button = ttk.Button(order_frame, text="Usuń zaznaczony", command=remove_select_item)
    button.grid(row=2, column=0, pady=(0, 10), sticky="ew")

    button = ttk.Button(order_frame, text="Zwiększ ilość", command=increase_quantity)
    button.grid(row=3, column=0, pady=(0, 10), sticky="ew")

    button = ttk.Button(order_frame, text="Zmniejsz ilość", command=reduce_quantity)
    button.grid(row=4, column=0, pady=(0, 10), sticky="ew")

    button = ttk.Button(order_frame, text="Wyszukaj", command=search)
    button.grid(row=5, column=0, pady=(0, 10), sticky="ew")

    button = ttk.Button(order_frame, text="Wygeneruj excel", command=generate_excel)
    button.grid(row=6, column=0, pady=(0, 10), sticky="ew")

    button = ttk.Button(order_frame, text="Przyjęcie produktów", command=admission)
    button.grid(row=7, column=0, pady=(0, 10), sticky="ew")

    button = ttk.Button(order_frame, text="Historia przyjęć", command=show_history)
    button.grid(row=8, column=0, pady=(0, 10), sticky="ew")


    inventory_frame = ttk.Frame(main_frame, padding=10)
    inventory_frame.grid(row=0, column=2, columnspan=8, padx=10, pady=10, sticky="nsew")

    data_entry_button = ttk.Button(inventory_frame, text="Pokaż stan magazynowy", command=populate_tree)
    data_entry_button.grid(row=0, column=1, pady=(0, 10), sticky="ew")

    tree = ttk.Treeview(inventory_frame)
    tree["columns"] = ("ID", "Nazwa", "Ilosc")
    tree.heading("ID", text="ID", anchor='center', command=lambda: sort_treeview(tree, "ID", False))
    tree.heading("Nazwa", text="Nazwa", anchor='center', command=lambda: sort_treeview2(tree, "Nazwa", False))
    tree.heading("Ilosc", text="Ilość", anchor='center', command=lambda: sort_treeview(tree, "Ilosc", False))
    tree.column("#0", width=0, minwidth=0, stretch=tk.NO)
    tree.column("ID", width=150, stretch=False)
    tree.column("Nazwa", width=150)
    tree.column("Ilosc", width=150)
    tree.grid(row=1, column=1, pady=(0, 10), sticky="nsew")

    add_to_order_button = ttk.Button(inventory_frame, text="Dodaj do zlecenia", command=add_to_order)
    add_to_order_button.grid(row=2, column=1, pady=(0, 10), sticky="ew", ipadx=20, ipady=10)

    order_tree = ttk.Treeview(inventory_frame)
    order_tree["columns"] = ("ID", "Nazwa", "Ilosc", "Numer")
    order_tree.heading("ID", text="ID", anchor='center')
    order_tree.heading("Nazwa", text="Nazwa", anchor='center')
    order_tree.heading("Ilosc", text="Ilość", anchor='center')
    order_tree.heading("Numer", text="Numer Zlecenia", anchor='center')
    order_tree.column("#0", width=0, minwidth=0, stretch=tk.NO)
    order_tree.column("ID", width=100, stretch=False)
    order_tree.column("Nazwa", width=150)
    order_tree.column("Ilosc", width=100)
    order_tree.column("Numer", width=100, stretch=False)
    order_tree.grid(row=3, column=1, pady=(0, 10), columnspan=2, sticky="nsew")

    remove_from_order_button = ttk.Button(inventory_frame, text="Usuń zaznaczony", command=remove_from_order)
    remove_from_order_button.grid(row=4, column=1, pady=(0, 10), sticky="ew")

    clear_order_button = ttk.Button(inventory_frame, text="Wyczyść listę", command=clear_window_without_send)
    clear_order_button.grid(row=5, column=1, pady=(0, 10), sticky="ew")

    save_to_db_button = ttk.Button(inventory_frame, text="Wyślij zlecenie", command=save_to_database)
    save_to_db_button.grid(row=6, column=1, pady=(0, 10), columnspan=2, sticky="ew", ipadx=20, ipady=10)


    order__frame = ttk.Frame(main_frame, padding=10)
    order__frame.grid(row=0, column=12, columnspan=8, padx=10, pady=10, sticky="nsew")

    data_entry_button = ttk.Button(order__frame, text="Pokaż zamówienia", command=populate_order_tree)
    data_entry_button.grid(row=0, column=0, pady=(0, 10), columnspan=2, sticky="ew")

    done_order_button = ttk.Button(order__frame, text="Zakończ zamówienie", command=remove_order_by_number)
    done_order_button.grid(row=2, column=0, pady=(0, 10), columnspan=2, sticky="ew", ipadx=20, ipady=10)

    done_order_button = ttk.Button(order__frame, text="Historia zamówień", command=history)
    done_order_button.grid(row=3, column=0, pady=(0, 10), columnspan=2, sticky="ew")

    order__tree = ttk.Treeview(order__frame)
    order__tree["columns"] = (
    "ID", "Nazwa", "Ilosc", "Numer", "DataDodania", "Zamawiajacy")
    order__tree.heading("ID", text="ID", anchor='center')
    order__tree.heading("Nazwa", text="Nazwa", anchor='center')
    order__tree.heading("Ilosc", text="Ilość", anchor='center')
    order__tree.heading("Numer", text="Numer", anchor='center')
    order__tree.heading("DataDodania", text="Data Dodania", anchor='center')
    order__tree.heading("Zamawiajacy", text="Zamawiajacy", anchor='center')

    order__tree.column("#0", width=0, minwidth=0, stretch=tk.NO)
    order__tree.column("ID", width=150, stretch=False)
    order__tree.column("Nazwa", width=150)
    order__tree.column("Ilosc", width=150)
    order__tree.column("Numer", width=150)
    order__tree.column("DataDodania", width=200)
    order__tree.column("Zamawiajacy", width=100)
    order__tree.grid(row=1, column=0, pady=(0, 10), columnspan=2, sticky="nsew")

    order__tree.heading("Numer", text="Numer", anchor='center',
                        command=lambda: sort_treeview(order__tree, "Numer", False))

    root.resizable(width=False, height=False)

    root.mainloop()


###################################### Pobieranie danych z bazy #####################################################


def data_entry2():
    def clear_treeview2():
        for item in order__tree.get_children():
            order__tree.delete(item)

    clear_treeview2()


def populate_order_tree():
    data_entry2()
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()
    cursor.execute("SELECT * from Zlecenie")
    rows = cursor.fetchall()
    for row in rows:
        order__tree.insert("", "end", values=row)
    conn.close()


def data_entry():
    def clear_treeview():
        for item in tree.get_children():
            tree.delete(item)

    clear_treeview()


def populate_tree():
    data_entry()
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()
    cursor.execute("SELECT * from Magazyn")
    rows = cursor.fetchall()
    for row in rows:
        tree.insert("", "end", values=row)
    conn.close()


def data_entry3():
    def clear_treeview3():
        for item in users_tree.get_children():
            users_tree.delete(item)

    clear_treeview3()

def populate_tree3():
    data_entry3()
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()
    cursor.execute("SELECT * from uzytkownicy")
    rows = cursor.fetchall()
    for row in rows:
        users_tree.insert("", "end", values=row)
    conn.close()


def data_entry4():
    def clear_treeview4():
        for item in history_tree.get_children():
            history_tree.delete(item)

    clear_treeview4()


def populate_tree4():
    data_entry4()
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()
    cursor.execute("SELECT * from Historia")
    rows = cursor.fetchall()
    for row in rows:
        history_tree.insert("", "end", values=row)
    conn.close()


def data_entry5():
    def clear_treeview5():
        for item in history_receipt_tree.get_children():
            history_receipt_tree.delete(item)

    clear_treeview5()


def populate_tree5():
    data_entry5()
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()
    cursor.execute("SELECT * from HistoriaPrzyjec")
    rows = cursor.fetchall()
    for row in rows:
        history_receipt_tree.insert("", "end", values=row)
    conn.close()


###################################### Podstawowe funkcjonalności magazynu #####################################################


def remove_select_item():
    select_item = tree.selection()
    if len(select_item) == 0:
        return

    select_itemm = select_item[0]

    date_select_item = tree.item(select_itemm, 'values')

    id_elementu = date_select_item[0]

    confirm = messagebox.askyesno("Potwierdzenie", "Czy na pewno chcesz usunąć ten element?")
    if confirm:
        if not is_product_used(id_elementu):
            remove(id_elementu)
            tree.delete(select_itemm)
        else:
            messagebox.showwarning("Ostrzeżenie", "Nie można usunąć produktu, ponieważ jest używany w tabeli Historia lub Zlecenia")

def is_product_used(product_id):
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM Zlecenie WHERE ID = %s", (product_id,))
    result_zlecenia = cursor.fetchone()

    cursor.execute("SELECT * FROM Historia WHERE ID = %s", (product_id,))
    result_historia = cursor.fetchone()

    conn.close()

    return result_historia or result_zlecenia

def remove(id):
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()

    cursor.execute("DELETE FROM Magazyn WHERE id=%s", (id,))

    conn.commit()
    conn.close()


def search():
    root = Toplevel()
    root.title("Wyszukiwanie po ID")

    root.grab_set()

    frame = ttk.Frame(root, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    label5 = ttk.Label(frame, text="Wyszukaj po ID:")
    label5.grid(row=0, column=0, sticky="w", pady=(20, 5))

    entry5 = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
    entry5.grid(row=0, column=1, padx=5, pady=(20, 5), ipady=10)

    def on_close():
        root.grab_release()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

    root.resizable(width=False, height=False)

    def search_id():
        id_to_search = entry5.get()
        conn = mysql.connector.connect(
            host="db4free.net",
            user="magazyn",
            password="Inzynierka",
            database="magazyn_pd"
        )
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM Magazyn WHERE id=%s", (id_to_search,))
        result = cursor.fetchone()

        conn.close()

        if result is not None:
            for item in tree.get_children():
                tree.delete(item)
            tree.insert("", "end", values=result)
        else:
            tk.messagebox.showinfo("Wynik wyszukiwania", "Brak rekordu o podanym ID.")

    button2 = ttk.Button(frame, text="Wyszukaj po ID", command=search_id)
    button2.grid(row=2, column=0, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)

    data_entry_button = ttk.Button(frame, text="Pokaż wszystkie", command=populate_tree)
    data_entry_button.grid(row=3, column=0, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)



def add_to_datebase():
    root = Toplevel()
    root.title("Dodawanie produktu do magazynu")

    root.grab_set()

    frame = ttk.Frame(root, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    label1 = ttk.Label(frame, text="ID:")
    label1.grid(row=0, column=0, sticky="w", pady=(20, 5))
    entry1 = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
    entry1.grid(row=0, column=1, padx=5, pady=(20, 5), ipady=10)

    label2 = ttk.Label(frame, text="Nazwa:")
    label2.grid(row=1, column=0, sticky="w", pady=(5, 10))
    entry2 = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
    entry2.grid(row=1, column=1, padx=5, pady=(5, 10), ipady=10)

    label3 = ttk.Label(frame, text="Ilość:")
    label3.grid(row=2, column=0, sticky="w", pady=(5, 10))
    entry3 = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
    entry3.grid(row=2, column=1, padx=5, pady=(5, 10), ipady=10)

    result_label = ttk.Label(frame, text="", foreground="red")
    result_label.grid(row=4, column=0, columnspan=2, pady=(15, 20))

    def on_close():
        root.grab_release()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

    root.resizable(width=False, height=False)

    def is_numeric(value):
        try:
            float(value)
            return True
        except ValueError:
            return False

    def check_and_add():
        id_value = entry1.get()
        nazwa_value = entry2.get()
        ilosc_value = entry3.get()

        if not is_numeric(id_value):
            tk.messagebox.showerror("Błąd", "Pole 'ID' musi zawierać wartość numeryczną.")
            return

        if not is_numeric(ilosc_value):
            tk.messagebox.showerror("Błąd", "Pole 'Ilość' musi zawierać wartość numeryczną.")
            return

        date = (id_value, nazwa_value, ilosc_value)

        if exists(id_value):
            tk.messagebox.showerror("Błąd", "Towar o podanym numerze magazynowym już istnieje.")
        else:
            add_to_datebase2(date)
            tk.messagebox.showinfo("Sukces", "Produkt został dodany.")

    def exists(value):
        conn = mysql.connector.connect(
            host="db4free.net",
            user="magazyn",
            password="Inzynierka",
            database="magazyn_pd"
        )
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Magazyn WHERE id=%s", (value,))
        result = cursor.fetchone()

        conn.close()

        return result is not None

    def add_to_datebase2(date):
        conn = mysql.connector.connect(
            host="db4free.net",
            user="magazyn",
            password="Inzynierka",
            database="magazyn_pd"
        )
        cursor = conn.cursor()

        cursor.execute("INSERT INTO Magazyn VALUES (%s, %s, %s)", date)

        conn.commit()
        conn.close()
        populate_tree()

    button = ttk.Button(frame, text="Dodaj do magazynu", command=check_and_add)
    button.grid(row=3, column=0, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)


def admission():

    def is_numeric(value):
        try:
            float(value)
            return True
        except ValueError:
            return False

    def populate_receipt_tree():
        for child in receipt_tree.get_children():
            receipt_tree.delete(child)
        for item in temp_data:
            receipt_tree.insert("", "end", values=item)

    def add_to_temp_data():
        id_value = entry1.get()
        name_value = entry2.get()
        quantity_value = entry3.get()

        if not is_numeric(id_value) or not is_numeric(quantity_value):
            messagebox.showerror("Błąd", "ID i Ilość muszą być wartościami numerycznymi.")
            return

        temp_data.append((id_value, name_value, quantity_value))
        populate_receipt_tree()
        entry1.delete(0, tk.END)
        entry2.delete(0, tk.END)
        entry3.delete(0, tk.END)

    def admission():
        for item in temp_data:
            id_value, name_value, quantity_value = item

            conn = mysql.connector.connect(
                host="db4free.net",
                user="magazyn",
                password="Inzynierka",
                database="magazyn_pd"
            )
            cursor = conn.cursor()

            cursor.execute("SELECT * FROM Magazyn WHERE id=%s", (id_value,))
            existing_row = cursor.fetchone()
            if existing_row:
                new_quantity = int(existing_row[2]) + int(quantity_value)
                cursor.execute("UPDATE Magazyn SET ilosc=%s WHERE id=%s", (new_quantity, id_value))
            else:
                date = (id_value, name_value, quantity_value)
                cursor.execute("INSERT INTO Magazyn VALUES (%s, %s, %s)", date)

            current_datetime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            history_date = (id_value, name_value, quantity_value, current_datetime, user_id)
            cursor.execute("INSERT INTO HistoriaPrzyjec VALUES (%s, %s, %s, %s, %s)", history_date)

            conn.commit()
            conn.close()

        messagebox.showinfo("Sukces", "Wszystkie produkty zostały przyjęte.")
        temp_data.clear()
        populate_receipt_tree()

    def import_from_excel():
        file_path = filedialog.askopenfilename(filetypes=[("Pliki Excela", "*.xlsx;*.xls")])

        if not file_path:
            return

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                id_value, name_value, quantity_value = map(str, row)

                if not is_numeric(id_value) or not is_numeric(quantity_value):
                    messagebox.showerror("Błąd", "ID i Ilość muszą być wartościami numerycznymi.")
                    return

                temp_data.append((id_value, name_value, quantity_value))
            populate_receipt_tree()
        except Exception as e:
            messagebox.showerror("Błąd", f"Wystąpił błąd podczas importu z Excela:\n{str(e)}")

    def remove_selected_item():
        selected_item = receipt_tree.selection()
        if selected_item:
            for item in selected_item:
                values = receipt_tree.item(item, 'values')
                temp_data.remove(tuple(values))
                receipt_tree.delete(item)

    root = Toplevel()
    root.title("Przyjęcie produktów")

    root.grab_set()

    frame = ttk.Frame(root, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    label1 = ttk.Label(frame, text="ID:")
    label1.grid(row=0, column=0, sticky="w", pady=(20, 5))
    entry1 = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
    entry1.grid(row=0, column=1, padx=5, pady=(20, 5), ipady=10)

    label2 = ttk.Label(frame, text="Nazwa:")
    label2.grid(row=1, column=0, sticky="w", pady=(5, 10))
    entry2 = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
    entry2.grid(row=1, column=1, padx=5, pady=(5, 10), ipady=10)

    label3 = ttk.Label(frame, text="Ilość:")
    label3.grid(row=2, column=0, sticky="w", pady=(5, 10))
    entry3 = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
    entry3.grid(row=2, column=1, padx=5, pady=(5, 10), ipady=10)

    result_label = ttk.Label(frame, text="", foreground="red")
    result_label.grid(row=5, column=0, columnspan=2, pady=(15, 20))

    button_add = ttk.Button(frame, text="Dodaj do tymczasowego przyjęcia", command=add_to_temp_data)
    button_add.grid(row=3, column=0, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)

    button_import = ttk.Button(frame, text="Import z Excela", command=import_from_excel)
    button_import.grid(row=4, column=0, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)

    button_remove = ttk.Button(frame, text="Usuń zaznaczone", command=remove_selected_item)
    button_remove.grid(row=7, column=0, columnspan=2)

    receipt_tree = ttk.Treeview(frame, columns=("ID", "Nazwa", "Ilość"))
    receipt_tree["show"] = "headings"
    receipt_tree.heading("ID", text="ID")
    receipt_tree.heading("Nazwa", text="Nazwa")
    receipt_tree.heading("Ilość", text="Ilość")
    receipt_tree.grid(row=6, column=0, columnspan=2, pady=(20, 15))

    button_accept = ttk.Button(frame, text="Przyjmij", command=admission)
    button_accept.grid(row=8, column=0, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)

    temp_data = []

    def on_close():
        root.grab_release()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

    root.resizable(width=False, height=False)

    root.mainloop()


def show_history():
    global history_receipt_tree

    root = Toplevel()
    root.title("Historia Przyjęć")

    root.grab_set()

    frame = ttk.Frame(root, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    history_receipt_tree = ttk.Treeview(frame, columns=("ID", "Nazwa", "Ilość", "Data Przyjęcia", "Przyjmujący"))
    history_receipt_tree["show"] = "headings"
    history_receipt_tree.heading("ID", text="ID")
    history_receipt_tree.heading("Nazwa", text="Nazwa")
    history_receipt_tree.heading("Ilość", text="Ilość")
    history_receipt_tree.heading("Data Przyjęcia", text="Data Przyjęcia")
    history_receipt_tree.heading("Przyjmujący", text="Przyjmujący")
    history_receipt_tree.grid(row=1, column=0, columnspan=2, pady=(20, 15))

    populate_tree5()

    def on_close():
        root.grab_release()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

    root.resizable(width=False, height=False)

    root.mainloop()


def increase_quantity():

    select_items = tree.selection()
    if len(select_items) == 0:

        return

    select_item = select_items[0]

    date_select_item = tree.item(select_item, 'values')

    id_item = date_select_item[0]

    increase_quantity_in_datebase(id_item)

    update_tree(select_item)

def increase_quantity_in_datebase(id):
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()

    cursor.execute("UPDATE Magazyn SET Ilosc = Ilosc + 1 WHERE id=%s", (id,))

    conn.commit()
    conn.close()

def update_tree(item):
    current_quantity = int(tree.item(item, 'values')[2])

    new_quantity = current_quantity + 1

    tree.item(item, values=(tree.item(item, 'values')[0], tree.item(item, 'values')[1], new_quantity))

def reduce_quantity():

    select_items = tree.selection()
    if len(select_items) == 0:
        return

    select_item = select_items[0]

    date_select_item = tree.item(select_item, 'values')

    id_item = date_select_item[0]

    reduce_quantity_in_datebase(id_item)

    update_tree2(select_item)


def reduce_quantity_in_datebase(id):
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()
    cursor.execute("UPDATE Magazyn SET Ilosc = Ilosc - 1 WHERE id=%s", (id,))

    conn.commit()
    conn.close()


def update_tree2(element):
    current_quantity = int(tree.item(element, 'values')[2])

    new_quantity = current_quantity - 1

    tree.item(element, values=(tree.item(element, 'values')[0], tree.item(element, 'values')[1], new_quantity))


def generate_excel():
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM Magazyn")
    data = cursor.fetchall()

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Pliki Excel", "*.xlsx")])

    if file_path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dane z bazy"

        headers = ["ID", "Nazwa", "Ilość"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)

        wb.save(file_path)

        conn.close()

        messagebox.showinfo("Sukces", f"Plik Excel został wygenerowany pomyślnie.")


def sort_treeview(tree, col, descending):
    data = [(int(tree.set(child, col)), child) for child in tree.get_children('')]
    data.sort(reverse=descending)
    for i, (val, child) in enumerate(data):
        tree.move(child, '', i)
    tree.heading(col, command=lambda: sort_treeview(tree, col, not descending))


def sort_treeview2(tree, col, descending):
    data = [(tree.set(child, col).lower(), child) for child in tree.get_children('')]
    data.sort(reverse=descending)
    for i, (val, child) in enumerate(data):
        tree.move(child, '', i)
    tree.heading(col, command=lambda: sort_treeview2(tree, col, not descending))


###################################### Tworzenie zlecenia #####################################################

current_order_number = 1

added_products = set()


def add_to_order():
    selected_items = tree.selection()

    for item_id in selected_items:
        item = tree.item(item_id, 'values')
        product_id = item[0]

        available_quantity = int(item[2])

        quantity_window = Toplevel()
        quantity_window.title("Ilość")

        quantity_window.grab_set()

        frame = ttk.Frame(quantity_window, padding=(30, 30, 30, 30))
        frame.grid(row=0, column=0, sticky="nsew")

        entry_label = ttk.Label(frame, text=f" {item[1]} (dostępna ilość: {available_quantity}):")
        entry_label.grid(row=0, column=0, sticky="w", pady=(20, 5))
        entry = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
        entry.grid(row=0, column=1, padx=5, pady=(20, 5), ipady=10)

        def get_last_order_number():
            conn = mysql.connector.connect(
                host="db4free.net",
                user="magazyn",
                password="Inzynierka",
                database="magazyn_pd"
            )
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(Numer) FROM Zlecenie")
            last_order_number = cursor.fetchone()[0]
            conn.close()
            return last_order_number if last_order_number is not None else 0

        def get_max_order_number_from_history():
            conn = mysql.connector.connect(
                host="db4free.net",
                user="magazyn",
                password="Inzynierka",
                database="magazyn_pd"
            )
            cursor = conn.cursor()
            cursor.execute("SELECT MAX(Numer) FROM Historia")
            max_order_number_from_history = cursor.fetchone()[0]
            conn.close()
            return max_order_number_from_history if max_order_number_from_history is not None else 0

        def confirm_quantity():
            nonlocal quantity_window
            quantity_str = entry.get()
            if quantity_str.isdigit():
                quantity = int(quantity_str)
                if 0 < quantity <= available_quantity:

                    reduce_quantity_in_database2(product_id, quantity)
                    populate_tree()

                    last_order_number = get_last_order_number()
                    max_order_number_from_history = get_max_order_number_from_history()

                    new_order_number = max(last_order_number, max_order_number_from_history) + 1

                    new_item = item[:2] + (quantity, new_order_number)
                    order_tree.insert('', 'end', values=new_item)

                    quantity_window.destroy()
                else:
                    tk.messagebox.showerror("Błąd",
                                            f"Podana ilość jest nieprawidłowa lub przekracza dostępną ilość ({available_quantity}).")

        confirm_button = ttk.Button(frame, text="Potwierdź", command=confirm_quantity)
        confirm_button.grid(row=1, column=0, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)

        def on_close():
            quantity_window.grab_release()
            quantity_window.destroy()

        quantity_window.protocol("WM_DELETE_WINDOW", on_close)

        quantity_window.resizable(width=False, height=False)

        quantity_window.mainloop()


def reduce_quantity_in_database2(id, quantity):
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()

    cursor.execute("UPDATE Magazyn SET Ilosc = Ilosc - %s WHERE id=%s", (quantity, id))

    conn.commit()
    conn.close()

removed_products = {}

def remove_from_order():
    selected_items = order_tree.selection()
    for item_id in selected_items:
        item = order_tree.item(item_id, 'values')
        product_id = item[0]
        quantity = item[2]

        if product_id in removed_products:
            removed_products[product_id] += quantity
        else:
            removed_products[product_id] = quantity

        order_tree.delete(item_id)

    restore_to_datebase()
    populate_tree()
def restore_to_datebase():
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()

    for product_id, quantity in removed_products.items():
        cursor.execute("UPDATE Magazyn SET Ilosc = Ilosc + %s WHERE id=%s", (quantity, product_id))

    conn.commit()
    conn.close()

    removed_products.clear()

def save_to_database():
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()

    current_datetime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    user_name = user_id

    for item in order_tree.get_children():
        values = order_tree.item(item, 'values')
        product_id, name, quantity, order_number = values

        cursor.execute(
            "INSERT INTO Zlecenie (ID, Nazwa, Ilosc, Numer, DataDodania, identyfikator) VALUES (%s, %s, %s, %s, %s, %s)",
            (product_id, name, quantity, order_number, current_datetime,
             user_name))

    conn.commit()
    conn.close()
    clear_window()


def clear_window():
    for item in order_tree.get_children():
        order_tree.delete(item)


def clear_window_without_send():

    for item_id in order_tree.get_children():
        item = order_tree.item(item_id, 'values')
        product_id = item[0]
        quantity = item[2]

        if product_id in removed_products:
            removed_products[product_id] += quantity
        else:
            removed_products[product_id] = quantity

    for item in order_tree.get_children():
        order_tree.delete(item)

    restore_to_datebase()
    populate_tree()


###################################### Potwierdzanie zlecenia #####################################################


def remove_order_by_number():
    root = Toplevel()
    root.title("Zakończenie zamówienia")

    root.grab_set()

    frame = ttk.Frame(root, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    entry_label = ttk.Label(frame, text="Numer zamówienia:")
    entry_label.grid(row=0, column=0, sticky="w", pady=(20, 5))
    entry = ttk.Entry(frame, width=50, font=('Arial', 12), justify='left')
    entry.grid(row=0, column=1, padx=5, pady=(20, 5), ipady=10)

    root.resizable(width=False, height=False)

    def on_close():
        root.grab_release()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

    root.resizable(width=False, height=False)

    def show_order_details(orders_data):
        details_text = ""
        for order_data in orders_data:
            details_text += f"Nazwa: {order_data[1]}\nIlość: {order_data[2]}\nNumer: {order_data[3]}\nData dodania: {order_data[4]}\nID_Uzytkownika: {order_data[5]}\n\n"

        result = messagebox.askokcancel("Potwierdzenie zamówienia",
                                        f"Czy na pewno chcesz przenieść te zamówienia do historii?\n\n{details_text}")
        return result

    def End():
        order_number = entry.get()

        if order_number.isdigit():
            order_number = int(order_number)

            conn = mysql.connector.connect(
                host="db4free.net",
                user="magazyn",
                password="Inzynierka",
                database="magazyn_pd"
            )
            cursor = conn.cursor()

            cursor.execute("SELECT * FROM Zlecenie WHERE Numer = %s", (order_number,))
            orders_data = cursor.fetchall()

            if show_order_details(orders_data):
                for order_data in orders_data:

                    insert_query = "INSERT INTO Historia (ID, Nazwa, Ilosc, Numer, DataDodania, Status, DataZakonczenia, Zamawiajacy, Realizujacy) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"

                    current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    history_data = (
                        order_data[0], order_data[1], order_data[2], order_data[3], order_data[4], 'Zakończone',
                        current_date, order_data[5], user_id)
                    cursor.execute(insert_query, history_data)

                    cursor.execute("DELETE FROM Zlecenie WHERE Numer = %s AND ID = %s", (order_number, order_data[0]))

                conn.commit()

                for item in order_tree.get_children():
                    order_tree.delete(item)

                populate_order_tree()

            conn.close()

    def Cancel():
        order_number = entry.get()

        if order_number.isdigit():
            order_number = int(order_number)

            conn = mysql.connector.connect(
                host="db4free.net",
                user="magazyn",
                password="Inzynierka",
                database="magazyn_pd"
            )
            cursor = conn.cursor()

            cursor.execute("SELECT * FROM Zlecenie WHERE Numer = %s", (order_number,))
            orders_data = cursor.fetchall()

            if show_order_details(orders_data):
                for order_data in orders_data:
                    insert_query = "INSERT INTO Historia (ID, Nazwa, Ilosc, Numer, DataDodania, Status, DataZakonczenia, Zamawiajacy, Realizujacy) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"

                    current_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    history_data = (
                        order_data[0], order_data[1], order_data[2], order_data[3], order_data[4], 'Anulowane',
                        current_date, order_data[5], user_id)
                    cursor.execute(insert_query, history_data)

                    cursor.execute("DELETE FROM Zlecenie WHERE Numer = %s AND ID = %s", (order_number, order_data[0]))

                conn.commit()

                for item in order_tree.get_children():
                    order_tree.delete(item)

                populate_order_tree()

            conn.close()

    remove_order_button = ttk.Button(frame, text="Potwierdź zamówienie", command=End)
    remove_order_button.grid(row=1, column=1, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)

    remove_order_button = ttk.Button(frame, text="Anuluj zamówienie", command=Cancel)
    remove_order_button.grid(row=2, column=1, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)



def history():
    global history_tree

    history_window = Toplevel()
    history_window.title("Historia Zleceń Magazynowych")

    history_window.grab_set()

    frame = ttk.Frame(history_window, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    history_tree = ttk.Treeview(frame, columns=(
        "ID", "Nazwa", "Ilosc", "Numer", "DataDodania", "Status", "DataZakonczenia", "Zamawiajacy", "Realizujacy"))
    history_tree["show"] = "headings"
    history_tree.heading("ID", text="ID", anchor='center')
    history_tree.heading("Nazwa", text="Nazwa", anchor='center')
    history_tree.heading("Ilosc", text="Ilość", anchor='center')
    history_tree.heading("Numer", text="Numer Zlecenia", anchor='center')
    history_tree.heading("DataDodania", text="Data Dodania", anchor='center')
    history_tree.heading("Status", text="Status", anchor='center')
    history_tree.heading("DataZakonczenia", text="Data Zakonczenia", anchor='center')
    history_tree.heading("Zamawiajacy", text="Zamawiajacy", anchor='center')
    history_tree.heading("Realizujacy", text="Realizujacy", anchor='center')
    history_tree.grid(row=1, column=0, columnspan=2, pady=(20, 15))

    populate_tree4()

    def on_close():
        history_window.grab_release()
        history_window.destroy()

    history_window.protocol("WM_DELETE_WINDOW", on_close)

    history_window.resizable(width=False, height=False)

    history_window.mainloop()


###################################### Funkcje administratora #####################################################

def hash_password(haslo):
    skrot = hashlib.sha256(haslo.encode()).hexdigest()
    return skrot

def user_edit1(id):

    def save_edit():
        new_name = name_entry.get()
        new_lastname = lastname_entry.get()
        new_level = level_var.get()
        new_email = email_entry.get()
        new_id = id_entry.get()

        conn = mysql.connector.connect(
            host="db4free.net",
            user="magazyn",
            password="Inzynierka",
            database="magazyn_pd"
        )
        cursor = conn.cursor()

        cursor.execute('SELECT imie, nazwisko, email FROM uzytkownicy WHERE id=%s', (id,))
        old=cursor.fetchone()
        old_name=old[0]
        old_lastname=old[1]
        old_email=old[2]

        if new_name != old_name or new_lastname != old_lastname:
            old_email = f"{new_name.lower()}.{new_lastname.lower()}@email.com"
        elif new_email != old_email:
            old_email = new_email

        cursor.execute('UPDATE uzytkownicy SET imie=%s, nazwisko=%s, rola=%s, email=%s, identyfikator=%s WHERE id=%s',
                       (new_name, new_lastname, new_level, old_email,new_id, id))

        conn.commit()

        populate_tree3()

        messagebox.showinfo("Sukces", "Edycja użytkownika zakończona pomyślnie.")

    root = Toplevel()
    root.title("Edycja użytkownika")

    root.grab_set()

    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()
    cursor.execute('SELECT imie, nazwisko, rola, email, identyfikator FROM uzytkownicy WHERE id=%s', (id,))
    date_user = cursor.fetchone()
    conn.close()

    frame = ttk.Frame(root, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    name_label = ttk.Label(frame, text="Imię:")
    name_label.grid(row=0, column=0, sticky="w", pady=(20, 5))
    name_entry = ttk.Entry(frame)
    name_entry.grid(row=0, column=1, pady=(20, 5), ipady=10)
    name_entry.insert(0, date_user[0])

    lastname_label = ttk.Label(frame, text="Nazwisko:")
    lastname_label.grid(row=1, column=0, sticky="w", pady=(5, 10))
    lastname_entry = ttk.Entry(frame)
    lastname_entry.grid(row=1, column=1, pady=(5, 10), ipady=10)
    lastname_entry.insert(0, date_user[1])

    level_label = ttk.Label(frame, text="Rola:")
    level_label.grid(row=2, column=0, sticky="w", pady=(5, 10))
    level = ["admin", "pracownik", "magazyn"]
    level_var = ttk.Combobox(frame, values=level)
    level_var.grid(row=2, column=1, pady=(5, 10), ipady=10)
    level_var.set(date_user[2])

    email_label = ttk.Label(frame, text="Email:")
    email_label.grid(row=3, column=0, sticky="w", pady=(5, 10))
    email_entry = ttk.Entry(frame)
    email_entry.grid(row=3, column=1, pady=(5, 10), ipady=10)
    email_entry.insert(0, date_user[3])

    id_label = ttk.Label(frame, text="Identfikator:")
    id_label.grid(row=4, column=0, sticky="w", pady=(5, 10))
    id_entry = ttk.Entry(frame)
    id_entry.grid(row=4, column=1, pady=(5, 10), ipady=10)
    id_entry.insert(0, date_user[4])

    save_button = ttk.Button(frame, text="Zapisz edycję", command=save_edit)
    save_button.grid(row=5, column=0, columnspan=2, pady=(10, 20))

    def on_close():
        root.grab_release()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

    root.resizable(width=False, height=False)

    root.mainloop()

def edit():
    select_items = users_tree.selection()
    if len(select_items) == 0:
        return

    root = Toplevel()
    root.title("Formularz edycji")

    root.grab_set()

    select_item = select_items[0]

    date_select_item = users_tree.item(select_item, 'values')

    id_item = date_select_item[0]

    frame = ttk.Frame(root, padding=30)
    frame.grid(row=0, column=0, sticky="nsew")

    create_button = ttk.Button(frame, text="Edytuj użytkownika", command=lambda: user_edit1(id_item))
    create_button.grid(row=1, column=0, pady=(0, 10), sticky="ew")

    create_button = ttk.Button(frame, text="Usuń użytkownika",
                               command=lambda: remove_select_item2(id_item, select_item))
    create_button.grid(row=1, column=1, pady=(0, 10), sticky="ew")

    create_button = ttk.Button(frame, text="Resetuj hasło użytkownika", command=lambda: reset_password(id_item))
    create_button.grid(row=1, column=3, pady=(0, 10), sticky="ew")

    def on_close():
        root.grab_release()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)

    root.resizable(width=False, height=False)

    root.mainloop()

def create():

    login_window = Toplevel()
    login_window.title("Formularz logowania")

    login_window.grab_set()

    frame = ttk.Frame(login_window, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    name_label = ttk.Label(frame, text="Imię:")
    name_label.grid(row=0, column=0, sticky="w", pady=(20, 5))
    name_entry = ttk.Entry(frame)
    name_entry.grid(row=0, column=1, pady=(20, 5), ipady=10)

    lastname_label = ttk.Label(frame, text="Nazwisko:")
    lastname_label.grid(row=1, column=0, sticky="w", pady=(5, 10))
    lastname_entry = ttk.Entry(frame)
    lastname_entry.grid(row=1, column=1, pady=(5, 10), ipady=10)

    password_label = ttk.Label(frame, text="Hasło:")
    password_label.grid(row=2, column=0, sticky="w", pady=(5, 10))
    password_entry = ttk.Entry(frame, show="*")
    password_entry.grid(row=2, column=1, pady=(5, 10), ipady=10)

    level_rola_label = ttk.Label(frame, text="Rola:")
    level_rola_label.grid(row=3, column=0, sticky="w", pady=(5, 10))

    level = ["admin", "pracownik", "magazyn"]
    level_rola_var = tk.StringVar()
    level_rola_var.set(level[0])
    level_menu = ttk.Combobox(frame, values=level, state="readonly")
    level_menu.grid(row=3, column=1, pady=(5, 10), ipady=10)

    result_label = ttk.Label(frame, text="")
    result_label.grid(row=6, column=0, columnspan=2, pady=(15, 20))

    def create_user():
        name = name_entry.get()
        lastname = lastname_entry.get()
        password = password_entry.get()
        new_level = level_menu.get()

        if not new_level:
            result_label.config(text="Wybierz rolę użytkownika.")
            return

        new_email = f"{name.lower()}.{lastname.lower()}@email.com"
        id = f"{new_level[0].upper()}_{name[0].upper()}{lastname[0].upper()}"

        conn = mysql.connector.connect(
            host="db4free.net",
            user="magazyn",
            password="Inzynierka",
            database="magazyn_pd"
        )
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM uzytkownicy WHERE email=%s', (new_email,))
        exists_user = cursor.fetchone()

        if exists_user:
            result_label.config(text="Użytkownik o podanym adresie email już istnieje.")
        else:
            cursor.execute(
                'INSERT INTO uzytkownicy (email, haslo, imie, nazwisko, rola, identyfikator) VALUES (%s, %s, %s, %s, %s, %s)',
                (new_email, hash_password(password), name, lastname, new_level, id))
            conn.commit()
            conn.close()
            populate_tree3()

            result_label.config(text="Nowy użytkownik został utworzony.")

            name_entry.delete(0, tk.END)
            lastname_entry.delete(0, tk.END)

    create_user_button = ttk.Button(frame, text="Utwórz użytkownika", command=create_user, width=20)
    create_user_button.grid(row=5, column=0, columnspan=2, pady=(20, 15), ipadx=20, ipady=10)

    login_window.resizable(width=False, height=False)

    login_window.mainloop()


def remove_select_item2(id_item, select_item):

    confirm = messagebox.askyesno("Potwierdzenie", "Czy na pewno chcesz usunąć ten element?")
    if confirm:
        remove_record2(id_item)

        users_tree.delete(select_item)

def remove_record2(id):
    conn = mysql.connector.connect(
        host="db4free.net",
        user="magazyn",
        password="Inzynierka",
        database="magazyn_pd"
    )
    cursor = conn.cursor()

    cursor.execute("DELETE FROM uzytkownicy WHERE id=%s", (id,))

    conn.commit()
    conn.close()


def on_close(root):
    root.destroy()
    show_login_screen()


def reset_password(id):
    def check_password2():
        new_password_1 = new_password_entry_1.get()
        new_password_2 = new_password_entry_2.get()

        if new_password_1 != new_password_2:
            messagebox.showerror("Błąd", "Podane hasła nie są identyczne.")
            return

        conn = mysql.connector.connect(
            host="db4free.net",
            user="magazyn",
            password="Inzynierka",
            database="magazyn_pd"
        )
        cursor = conn.cursor()

        cursor.execute('UPDATE uzytkownicy SET haslo=%s WHERE id=%s', (hash_password(new_password_1), id))
        conn.commit()

        messagebox.showinfo("Sukces", "Hasło zostało pomyślnie zmienione.")

    password_reset_window = Toplevel()
    password_reset_window.title("Formularz resetowania hasła")

    password_reset_window.grab_set()

    frame = ttk.Frame(password_reset_window, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    new_password_label_1 = ttk.Label(frame, text="Nowe hasło:")
    new_password_label_1.grid(row=0, column=0, sticky="w", pady=(20, 5))
    new_password_entry_1 = ttk.Entry(frame, show="*")
    new_password_entry_1.grid(row=0, column=1, pady=(20, 5), ipady=10)

    new_password_label_2 = ttk.Label(frame, text="Potwierdź nowe hasło:")
    new_password_label_2.grid(row=1, column=0, sticky="w", pady=(5, 10))
    new_password_entry_2 = ttk.Entry(frame, show="*")
    new_password_entry_2.grid(row=1, column=1, pady=(5, 10), ipady=10)

    change_password_button = ttk.Button(frame, text="Zmień hasło", command=check_password2)
    change_password_button.grid(row=2, column=0, columnspan=2, pady=(10, 20))

    def on_close():
        password_reset_window.grab_release()
        password_reset_window.destroy()

    password_reset_window.protocol("WM_DELETE_WINDOW", on_close)

    password_reset_window.resizable(width=False, height=False)

    password_reset_window.mainloop()

def change_password():

    def check_password():
        old_password = old_password_entry.get()
        new_password = new_password_entry.get()

        conn = mysql.connector.connect(
            host="db4free.net",
            user="magazyn",
            password="Inzynierka",
            database="magazyn_pd"
        )
        cursor = conn.cursor()
        cursor.execute('SELECT haslo FROM uzytkownicy WHERE identyfikator=%s', (user_id,))
        result = cursor.fetchone()

        if result:
            password_in_datebase = result[0]
            hash = hash_password(old_password)

            if password_in_datebase == hash:
                cursor.execute('UPDATE uzytkownicy SET haslo=%s WHERE identyfikator=%s',
                               (hash_password(new_password), user_id))
                conn.commit()
                messagebox.showinfo("Sukces", "Hasło zostało pomyślnie zmienione.")
            else:
                messagebox.showerror("Błąd", "Podane stare hasło jest nieprawidłowe.")
        else:
            messagebox.showerror("Błąd", "Wystąpił błąd podczas pobierania hasła z bazy danych.")

    password_change_window = Toplevel()
    password_change_window.title("Formularz zmiany hasła")

    password_change_window.grab_set()

    frame = ttk.Frame(password_change_window, padding=(30, 30, 30, 30))
    frame.grid(row=0, column=0, sticky="nsew")

    old_password_label = ttk.Label(frame, text="Stare hasło:")
    old_password_label.grid(row=0, column=0, sticky="w", pady=(20, 5))
    old_password_entry = ttk.Entry(frame, show="*")
    old_password_entry.grid(row=0, column=1, pady=(20, 5), ipady=10)

    new_password_label = ttk.Label(frame, text="Nowe hasło:")
    new_password_label.grid(row=1, column=0, sticky="w", pady=(5, 10))
    new_password_entry = ttk.Entry(frame, show="*")
    new_password_entry.grid(row=1, column=1, pady=(5, 10), ipady=10)

    change_password_button = ttk.Button(frame, text="Zmień hasło", command=check_password)
    change_password_button.grid(row=2, column=0, columnspan=2, pady=(10, 20))

    def on_close():
        password_change_window.grab_release()
        password_change_window.destroy()

    password_change_window.protocol("WM_DELETE_WINDOW", on_close)

    password_change_window.resizable(width=False, height=False)

    password_change_window.mainloop()

show_login_screen()
